import base64
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

OUTPUT_DIR = Path("data/output")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def _auto_width(ws):
    """Best-effort column width adjustment."""
    for col_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 35)


def _to_base64(img_path: Path) -> str:
    with open(img_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


def create_charts(clean_df: pd.DataFrame, group_col: Optional[str], numeric_cols: List[str]):
    """Create multiple matplotlib charts and return list of dicts with title, path, and base64."""
    charts = []
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if not numeric_cols:
        return charts

    first_num = numeric_cols[0]
    second_num = numeric_cols[1] if len(numeric_cols) > 1 else None

    # 1) Distribution
    plt.figure(figsize=(6, 4))
    clean_df[first_num].dropna().hist(bins=20, color="#4f46e5")
    plt.title(f"Distribution of {first_num}")
    plt.xlabel(first_num)
    plt.ylabel("Frequency")
    dist_path = OUTPUT_DIR / f"dist_{first_num}_{timestamp}.png"
    plt.tight_layout()
    plt.savefig(dist_path)
    plt.close()
    charts.append({"title": "Distribution", "path": str(dist_path), "image_base64": _to_base64(dist_path)})

    # 2) Boxplot
    plt.figure(figsize=(4, 4))
    clean_df[first_num].plot(kind="box", vert=True)
    plt.title(f"Boxplot of {first_num}")
    box_path = OUTPUT_DIR / f"box_{first_num}_{timestamp}.png"
    plt.tight_layout()
    plt.savefig(box_path)
    plt.close()
    charts.append({"title": "Boxplot", "path": str(box_path), "image_base64": _to_base64(box_path)})

    # 3) Grouped bar (mean)
    if group_col and group_col in clean_df.columns:
        plt.figure(figsize=(7, 4))
        grouped = clean_df.groupby(group_col)[first_num].mean().sort_values(ascending=False).head(12)
        grouped.plot(kind="bar", color="#10b981")
        plt.title(f"Average {first_num} by {group_col}")
        plt.xlabel(group_col)
        plt.ylabel(f"Average {first_num}")
        plt.xticks(rotation=45, ha="right")
        bar_path = OUTPUT_DIR / f"group_{group_col}_{timestamp}.png"
        plt.tight_layout()
        plt.savefig(bar_path)
        plt.close()
        charts.append({"title": "Grouped Average", "path": str(bar_path), "image_base64": _to_base64(bar_path)})

    # 4) Pie of top category
    cat_cols = [c for c in clean_df.columns if c not in numeric_cols and clean_df[c].nunique() <= 20]
    if cat_cols:
        top_cat = cat_cols[0]
        plt.figure(figsize=(5, 5))
        clean_df[top_cat].value_counts().head(8).plot(kind="pie", autopct="%1.0f%%", pctdistance=0.85)
        plt.title(f"Top {top_cat}")
        pie_path = OUTPUT_DIR / f"pie_{top_cat}_{timestamp}.png"
        plt.tight_layout()
        plt.savefig(pie_path)
        plt.close()
        charts.append({"title": f"Top {top_cat}", "path": str(pie_path), "image_base64": _to_base64(pie_path)})

    # 5) Line by date (if datetime exists)
    date_cols = clean_df.select_dtypes(include=["datetime64[ns]"]).columns.tolist()
    if date_cols:
        dt_col = date_cols[0]
        plt.figure(figsize=(7, 4))
        ts = clean_df.set_index(dt_col)[first_num].resample("D").mean().dropna().head(60)
        ts.plot(color="#f97316")
        plt.title(f"{first_num} over time ({dt_col})")
        plt.xlabel("Date")
        plt.ylabel(first_num)
        line_path = OUTPUT_DIR / f"line_{dt_col}_{timestamp}.png"
        plt.tight_layout()
        plt.savefig(line_path)
        plt.close()
        charts.append({"title": f"{first_num} over time", "path": str(line_path), "image_base64": _to_base64(line_path)})

    # 6) Scatter of first two numerics
    if second_num:
        plt.figure(figsize=(6, 4))
        plt.scatter(clean_df[first_num], clean_df[second_num], alpha=0.6, color="#6366f1")
        plt.xlabel(first_num)
        plt.ylabel(second_num)
        plt.title(f"{first_num} vs {second_num}")
        scatter_path = OUTPUT_DIR / f"scatter_{first_num}_{second_num}_{timestamp}.png"
        plt.tight_layout()
        plt.savefig(scatter_path)
        plt.close()
        charts.append({"title": f"{first_num} vs {second_num}", "path": str(scatter_path), "image_base64": _to_base64(scatter_path)})

    # 7) Correlation heatmap (bonus)
    if len(numeric_cols) >= 2:
        plt.figure(figsize=(6, 5))
        corr = clean_df[numeric_cols].corr()
        plt.imshow(corr, cmap="coolwarm", interpolation="nearest")
        plt.colorbar()
        plt.xticks(range(len(numeric_cols)), numeric_cols, rotation=45, ha="right")
        plt.yticks(range(len(numeric_cols)), numeric_cols)
        plt.title("Correlation Heatmap")
        heat_path = OUTPUT_DIR / f"heatmap_{timestamp}.png"
        plt.tight_layout()
        plt.savefig(heat_path)
        plt.close()
        charts.append({"title": "Correlation Heatmap", "path": str(heat_path), "image_base64": _to_base64(heat_path)})

    return charts


def generate_excel_report(
    cleaned_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    grouped_df: pd.DataFrame,
    charts,
    group_col: Optional[str],
) -> Path:
    """Create a multi-sheet Excel report with embedded charts; return path."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"analysis_report_{timestamp}.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        cleaned_df.to_excel(writer, sheet_name="Cleaned Data", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        if not grouped_df.empty:
            grouped_df.to_excel(writer, sheet_name="Grouped", index=False)

        workbook = writer.book
        # Add chart sheet
        chart_sheet = workbook.create_sheet("Charts")
        row_anchor = 1
        for chart in charts:
            title = chart.get("title", "Chart")
            img_path = chart.get("path")
            try:
                img = XLImage(img_path)
                img.anchor = f"A{row_anchor}"
                chart_sheet.add_image(img)
                chart_sheet.cell(row=row_anchor, column=1, value=title)
                row_anchor += int(img.height / 20) + 3  # heuristic spacing
            except Exception:
                logger.exception("Failed to embed chart %s", img_path)

        # autosize columns
        for ws in workbook.worksheets:
            _auto_width(ws)

    logger.info("Report written to %s", output_path)
    return output_path
